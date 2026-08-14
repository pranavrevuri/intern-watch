"""
Scrapes extern.com's per-company internship guides for expected summer
2027 application-open windows, career page links, and program facts,
then maintains two artifacts in the repo:

  extern_calendar.json  - machine-readable source of truth (diffed
                          between runs so email alerts fire only on
                          real changes)
  InternWatch.xlsx      - human-readable watchlist regenerated every
                          run, sorted so whatever opens soonest is at
                          the top

Every Extern guide carries a "Quick Facts" HTML table with labeled
rows (Where to apply / Application window (2027-28) / Rolling? / ...)
plus a FAQPage JSON-LD block. Both are server-rendered, so plain HTTP
is enough - no headless browser, no LLM calls, no per-run cost.

Extern's dates are PROJECTIONS from prior cycles, refreshed roughly
monthly, which is why this runs weekly. Live "posting just went up"
detection stays monitor.py's job.

The companies scraped live in extern_companies.json (tech guides
only). New guides Extern publishes are detected via their sitemap and
listed in the email for manual review - they may be non-tech, so they
are never auto-added.

Run manually with:  python extern_watch.py
"""

import json
import os
import re
import sys
import time
import html as html_lib
import urllib.request
from datetime import datetime, timezone, date

CONFIG_FILE = "extern_companies.json"
CALENDAR_FILE = "extern_calendar.json"
XLSX_FILE = "InternWatch.xlsx"
GUIDE_URL = "https://www.extern.com/post/{slug}-internship-guide"
SITEMAP_URL = "https://www.extern.com/sitemap.xml"
FETCH_DELAY_SECS = 0.5

MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "oct": 10, "nov": 11, "dec": 12,
}
# Seasons map to a representative open window (northern hemisphere).
SEASONS = {"spring": (3, 5), "summer": (6, 8), "fall": (9, 11), "autumn": (9, 11), "winter": (12, 2)}

# Phrases in the "Application window" fact meaning there is no cohort
# cycle to put on a calendar - the company hires ad hoc / year-round.
# Only checked against the first sentence: later sentences often
# mention "rolling"/"year-round" while the first names a real window.
NO_CYCLE_PATTERNS = [
    r"no cohort", r"no fixed window", r"no recurring", r"year[- ]round",
    r"no structured", r"ad hoc", r"no set (?:window|cycle)",
    r"no internship cycle", r"no dedicated [^.]*(?:window|cycle)",
]


def fetch(url, timeout=30):
    req = urllib.request.Request(url, headers={
        "User-Agent": "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/126.0 Safari/537.36"
    })
    with urllib.request.urlopen(req, timeout=timeout) as resp:
        return resp.read().decode("utf-8", errors="replace")


def strip_tags(fragment):
    text = re.sub(r"<[^>]+>", " ", fragment)
    return re.sub(r"\s+", " ", html_lib.unescape(text)).strip()


def parse_quick_facts(raw):
    """The Quick Facts table right after the <h2>Quick Facts</h2>
    heading: two columns, Fact | Detail. Returns {fact: (text, [hrefs])}
    keyed by lowercased fact label."""
    m = re.search(r"Quick Facts</h2>(.*?)</table>", raw, re.S)
    if not m:
        return {}
    facts = {}
    for row in re.findall(r"<tr[^>]*>(.*?)</tr>", m.group(1), re.S):
        cells = re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", row, re.S)
        if len(cells) < 2:
            continue
        key = strip_tags(cells[0]).lower()
        links = re.findall(r'href="(https?://[^"]+)"', cells[1])
        facts[key] = (strip_tags(cells[1]), links)
    return facts


def get_fact(facts, *prefixes):
    """Fact labels vary slightly between guides ('Visa sponsorship' vs
    'Visa sponsorship / return offers'), so match by prefix."""
    for key, val in facts.items():
        if any(key.startswith(p) for p in prefixes):
            return val
    return None


def parse_faqs(raw):
    """All (question, answer) pairs from the guide's FAQPage JSON-LD.
    The questions are templated across guides ('When do X internship
    applications open...', 'What is the X internship interview process
    like?', 'Does X sponsor visas for interns?'), which makes them the
    most reliable source for those fields."""
    faqs = []
    for m in re.finditer(r'<script type="application/ld\+json"[^>]*>(.*?)</script>', raw, re.S):
        try:
            data = json.loads(m.group(1))
        except (ValueError, TypeError):
            continue
        for item in data if isinstance(data, list) else [data]:
            if not isinstance(item, dict) or item.get("@type") != "FAQPage":
                continue
            for q in item.get("mainEntity", []):
                faqs.append((q.get("name") or "",
                             strip_tags(q.get("acceptedAnswer", {}).get("text", ""))))
    return faqs


def faq_answer(faqs, pattern):
    for q, a in faqs:
        if re.search(pattern, q, re.I):
            return a
    return None


# A date preceded by "for" or followed by a cohort noun names the
# internship itself ("for Dec 2027 start", "the Summer 2027 SWE
# internship"), never the application window.
COHORT_BEFORE = re.compile(r"\bfor\s+(?:an?\s+|the\s+)?$")
COHORT_AFTER = re.compile(r"^[\s,]*(?:start|cohort|intern(?:ship)?s?|class|swe|roles?|target|postings?)\b", re.I)


def _clause_tokens(clause):
    """(month, year_or_None, year_was_explicit) tokens from one clause:
    month names, 'early/mid/late YYYY' period phrases, and season+year
    phrases (mapped to the season's first month)."""
    # The optional day-of-month must not swallow the first digits of a
    # 4-digit year ("oct 2026" is month+year, "sep 17, 2025" has a day).
    month_pat = r"\b(jan|feb|mar|apr|may|jun|jul|aug|sep|oct|nov|dec)[a-z]*\.?(?:\s+~?\d{1,2}(?:st|nd|rd|th)?(?!\d),?)?\s*(\d{4})?"
    period_pat = r"\b(early|mid|late)[-\s]+(\d{4})\b"
    season_pat = r"\b(spring|summer|fall|autumn|winter)\s+(\d{4})\b"
    PERIOD_MONTH = {"early": 2, "mid": 7, "late": 11}
    matches = []
    for tm in re.finditer(month_pat, clause):
        matches.append((tm, MONTHS[tm.group(1)], int(tm.group(2)) if tm.group(2) else None))
    for tm in re.finditer(period_pat, clause):
        # "late July 2026" is already a month token; only take period
        # phrases where early/mid/late attaches straight to a year.
        matches.append((tm, PERIOD_MONTH[tm.group(1)], int(tm.group(2))))
    for tm in re.finditer(season_pat, clause):
        matches.append((tm, SEASONS[tm.group(1)][0], int(tm.group(2))))
    tokens = []
    for tm, mo, yr in sorted(matches, key=lambda x: x[0].start()):
        if COHORT_BEFORE.search(clause[max(0, tm.start() - 12):tm.start()]):
            continue
        if COHORT_AFTER.search(clause[tm.end():tm.end() + 16]):
            continue
        tokens.append([mo, yr, yr is not None])
    return tokens


def _resolve_years(tokens, today):
    """Months often share one trailing year ("Aug to Oct 2026"): borrow
    the nearest explicit year, bumping a borrowed year forward when it
    would land before the window start ("September-January" spans the
    new year). A clause with no year at all ("expected September")
    means the nearest upcoming occurrence of that month."""
    years = [t[1] for t in tokens]
    for i, tok in enumerate(tokens):
        if tok[1] is None:
            later = next((y for y in years[i:] if y), None)
            earlier = next((y for y in reversed(years[:i]) if y), None)
            tok[1] = later or earlier
            if tok[1] is None:
                tok[1] = today.year if tok[0] >= today.month - 3 else today.year + 1
    for tok in tokens[1:]:
        if not tok[2] and (tok[1], tok[0]) < (tokens[0][1], tokens[0][0]):
            tok[1] += 1
    return tokens


def parse_window(text, today=None):
    """Extract an (open_start, open_end) month window like
    ('2026-08', '2026-10') from Extern's prose, or None if no dated
    window is stated.

    The prose routinely mentions dates that are NOT the open window -
    prior cycles ("posted ~Sep 17, 2025"), deadlines ("close by early
    Feb 2027"), and the cohort itself ("for summer 2027") - so parsing
    works through candidate clauses in falling order of trust, drops
    dates outside the plausible range for the current cycle (prior
    cycles sit in the past, next cycles far ahead), and the first
    clause yielding dates wins."""
    if not text:
        return None
    today = today or date.today()
    low = text.lower()
    first_sentence = low.split(".")[0]
    if any(re.search(p, first_sentence) for p in NO_CYCLE_PATTERNS):
        return None

    candidates = []
    m = re.search(r"expect\w*[^.;]{0,60}?to (?:open|appear|post|go live|launch)([^.;]*)", low)
    if m:
        candidates.append(m.group(1))
    m = re.search(r"\b(?:open(?:s|ed|ing)?|went live|goes? live)\b([^.;]*)", low)
    if m:
        candidates.append(m.group(1))
    m = re.search(r"\bexpect(?:ed)?\b([^.;]*)", low)
    if m:
        candidates.append(m.group(1))
    candidates.append(low)

    lo_month = today.year * 12 + today.month - 3
    hi_month = today.year * 12 + today.month + 16
    for clause in candidates:
        # Deadline/close phrasing inside the clause is about the END of
        # applications, not the open window - cut it off.
        clause = re.sub(r",?\s*rolling (?:to|until|through)[^.;]*", "", clause)
        clause = re.sub(r"[,;]?\s*(?:and|with)?\s*(?:deadlines?|to close|closes?|closing)\b[^.;]*", "", clause)

        tokens = _resolve_years(_clause_tokens(clause), today)
        tokens = [t for t in tokens if lo_month <= t[1] * 12 + t[0] <= hi_month]
        if not tokens:
            continue
        start = tokens[0]
        # The window end is the furthest date within 10 months of the
        # start - far enough for "Sep 2026, runs through Feb 2027",
        # short enough to skip unrelated later dates.
        horizon = start[1] * 12 + start[0] + 10
        in_range = [t for t in tokens if start[1] * 12 + start[0] <= t[1] * 12 + t[0] <= horizon]
        end = max(in_range, key=lambda t: (t[1], t[0])) if in_range else start
        return (f"{start[1]:04d}-{start[0]:02d}", f"{end[1]:04d}-{end[0]:02d}")
    return None


def classify(entry, today):
    """Bucket used for sorting and the Status column. Projected
    windows, so 'open' means 'Extern expects it to be open - go
    check', not 'a posting is confirmed live'."""
    if entry.get("no_formal_program"):
        return "no_program"
    # Extern sometimes states outright that roles are already live
    # ("OPEN NOW (as of July 2026)", "open until filled") - believe
    # that over a parsed window whose start month has passed.
    if re.search(r"open now|already open|currently open|open until filled",
                 entry.get("window_text") or "", re.I):
        return "in_window"
    window = entry.get("window")
    if not window:
        return "continuous" if entry.get("window_text") else "unknown"
    start, end = window
    cur = f"{today.year:04d}-{today.month:02d}"
    if cur < start:
        start_dt = date(int(start[:4]), int(start[5:7]), 1)
        return "upcoming" if (start_dt - today).days <= 62 else "later"
    if cur > end:
        # One month of grace: rolling reviews routinely run past the
        # posted window, so "ended last month" still means check now.
        cur_m = today.year * 12 + today.month
        end_m = int(end[:4]) * 12 + int(end[5:7])
        return "in_window" if cur_m - end_m == 1 else "passed"
    return "in_window"


STATUS_LABELS = {
    "in_window": "Window open - check now",
    "upcoming": "Opening soon",
    "later": "Later this cycle",
    "continuous": "No fixed cycle - monitor",
    "passed": "Window likely passed",
    "no_program": "No formal program",
    "unknown": "Unclear - see guide",
    "fetch_failed": "Fetch failed",
}
STATUS_ORDER = ["in_window", "upcoming", "later", "continuous", "passed", "unknown", "fetch_failed", "no_program"]


# Fact-row links that are citations, not places to apply.
NON_CAREER_LINK = re.compile(r"glassdoor|levels\.fyi|linkedin\.com|reddit\.com|teamblind|blind\.com|indeed\.com|extern\.com")

# Track names that count as tech for the watchlist. Deliberately close
# to monitor.py's ROLE_KEYWORDS; "Digital Marketing"-style tracks are
# excluded by the veto even though they contain a match.
TECH_TRACK = re.compile(
    r"tech(?:nology)?|software|\bswe\b|\bsde\b|engineer|data\s+(?:scien|analyt|engineer)|"
    r"machine\s+learning|\bml\b|\bai\b|artificial intelligence|cyber|"
    r"information\s+(?:technology|security)|computer|quant|cloud|devops|analytics|digital|"
    r"\bstep\b|applied scien|research scientist|student researcher",
    re.I)
TECH_TRACK_VETO = re.compile(r"marketing|sales|brand|communicat|public relations|\bhr\b|people|recruit", re.I)


def parse_tracks(raw):
    """Track/program names from the guide's tracks table (the one
    headed 'Track' or 'Program' in the 'Which programs should you
    target' section). Empty list when a guide has no such table."""
    tracks = []
    for t in re.findall(r"<table>(.*?)</table>", raw, re.S):
        rows = re.findall(r"<tr[^>]*>(.*?)</tr>", t, re.S)
        cells = [re.findall(r"<t[dh][^>]*>(.*?)</t[dh]>", r, re.S) for r in rows]
        cells = [[strip_tags(c) for c in row] for row in cells if row]
        if not cells or not cells[0]:
            continue
        if not re.match(r"^(track|program|role|internship)", cells[0][0], re.I):
            continue
        tracks.extend(row[0][:80] for row in cells[1:] if row and row[0])
    return tracks


def tech_tracks_of(tracks):
    out = []
    for t in tracks:
        if t in out or not TECH_TRACK.search(t):
            continue
        if TECH_TRACK_VETO.search(t) and not re.search(r"tech|software|engineer|data|cyber|computer", t, re.I):
            continue
        out.append(t)
    return out


def parse_guide(raw, name, slug, today=None):
    facts = parse_quick_facts(raw)
    entry = {"name": name, "slug": slug, "guide_url": GUIDE_URL.format(slug=slug)}

    where = get_fact(facts, "where to apply")
    entry["career_urls"] = (where[1][:2] if where else []) or []
    if not entry["career_urls"]:
        # Guides without a "Where to apply" row (no-program companies
        # mostly) still link the careers page from some other row.
        fallback = [u for _, (_, links) in sorted(facts.items()) for u in links
                    if not NON_CAREER_LINK.search(u)]
        entry["career_urls"] = fallback[:2]
    if where and not entry["career_urls"]:
        entry["career_note"] = where[0][:200]

    window = get_fact(facts, "application window")
    window_text_full = window[0] if window else None

    rolling = get_fact(facts, "rolling")
    entry["rolling"] = rolling[0][:160] if rolling else None

    pay = get_fact(facts, "compensation")
    entry["pay"] = pay[0][:160] if pay else None

    programs = get_fact(facts, "# programs", "programs", "# tracks", "tracks")
    programs_text = programs[0] if programs else ""
    # What you'd actually be applying for - a row is never just "a
    # company" but named tracks ("Technology", "STEP, SWE, Student
    # Researcher..."), taken from the guide's tracks table when it has
    # one, else the Quick Facts programs row.
    tracks = parse_tracks(raw)
    tech = tech_tracks_of(tracks)
    entry["tech_tracks"] = "; ".join(tech)[:280] if tech else None
    entry["tracks_total"] = len(tracks)
    entry["programs"] = programs_text[:280] or None
    formal = get_fact(facts, "formal internship", "internship program", "current status")
    formal_text = formal[0] if formal else ""
    title_m = re.search(r"<title>(.*?)</title>", raw, re.S)
    title = strip_tags(title_m.group(1)) if title_m else ""
    entry["no_formal_program"] = bool(
        re.search(r"\b(?:0|no)\s+formal\b", programs_text, re.I)
        or re.search(r"^No[.,\s]", formal_text)
        or re.search(r"no formal (?:intern(?:ship)?\s+)?(?:program|cycle)", (window_text_full or "") + title, re.I)
        or re.search(r"no internships? posted", formal_text, re.I)
    )

    faqs = parse_faqs(raw)
    entry["faq_open_answer"] = faq_answer(faqs, r"when do .{0,60}open")
    # The interview pipeline (OA -> phone -> loop/superday) and the
    # visa/work-auth stance, for their watchlist columns. Visa prefers
    # the Quick Facts row; the FAQ is the fallback.
    interview = faq_answer(faqs, r"interview process|interview like")
    if not interview:
        # Older guides have no FAQ block; fall back to the standard
        # "What Is the X Application and Interview Process Like?" section.
        m = re.search(r"Application and Interview Process[^<]*</h2>(.*?)<h[23]", raw, re.S)
        interview = strip_tags(m.group(1)) if m else None
    entry["interview"] = (interview or "")[:300] or None
    visa = get_fact(facts, "visa")
    entry["visa"] = ((visa[0] if visa else None)
                     or faq_answer(faqs, r"visa|sponsor|international student") or "")[:220] or None
    # Parse before truncating for storage - the dated clause can sit
    # past the storage cap.
    entry["window"] = (parse_window(window_text_full, today)
                       or parse_window(entry["faq_open_answer"], today))
    entry["window_text"] = window_text_full[:400] if window_text_full else None
    if entry["faq_open_answer"]:
        entry["faq_open_answer"] = entry["faq_open_answer"][:400]
    return entry


def scrape_all(companies, log=print):
    entries = []
    for i, comp in enumerate(companies):
        slug = comp["slug"]
        try:
            raw = fetch(GUIDE_URL.format(slug=slug))
            entry = parse_guide(raw, comp["name"], slug)
        except Exception as exc:  # noqa: BLE001 - one bad page must not kill the run
            log(f"  ! {comp['name']}: {exc}")
            entry = {"name": comp["name"], "slug": slug,
                     "guide_url": GUIDE_URL.format(slug=slug), "fetch_error": str(exc)}
        entries.append(entry)
        log(f"[{i + 1}/{len(companies)}] {comp['name']}: "
            f"{(entry.get('window') or entry.get('window_text') or 'n/a') if not entry.get('fetch_error') else 'FETCH ERROR'}")
        time.sleep(FETCH_DELAY_SECS)
    return entries


def sitemap_guide_slugs():
    raw = fetch(SITEMAP_URL)
    locs = re.findall(r"<loc>https://www\.extern\.com/post/([a-z0-9-]+)-internship-guide</loc>", raw)
    return sorted(set(locs))


def diff_calendars(old_entries, new_entries):
    """Human-readable change lines; empty list means no email."""
    old = {e["slug"]: e for e in old_entries}
    new = {e["slug"]: e for e in new_entries}
    changes = []
    for slug, e in new.items():
        o = old.get(slug)
        if o is None:
            changes.append(f"NEW COMPANY: {e['name']} - {e.get('window_text') or 'see guide'}")
            continue
        if e.get("fetch_error") or o.get("fetch_error"):
            continue  # transient errors are not calendar changes
        if e.get("window_text") != o.get("window_text"):
            changes.append(
                f"{e['name']}: application window changed\n"
                f"    was: {o.get('window_text')}\n"
                f"    now: {e.get('window_text')}"
            )
        if e.get("career_urls") != o.get("career_urls"):
            changes.append(f"{e['name']}: career links changed to {', '.join(e.get('career_urls') or ['(none)'])}")
        if e.get("no_formal_program") != o.get("no_formal_program"):
            started = "now HAS a formal program" if not e["no_formal_program"] else "no longer lists a formal program"
            changes.append(f"{e['name']}: {started}")
        if e.get("tech_tracks") != o.get("tech_tracks"):
            changes.append(
                f"{e['name']}: tech tracks changed\n"
                f"    was: {o.get('tech_tracks')}\n"
                f"    now: {e.get('tech_tracks')}"
            )
    for slug, o in old.items():
        if slug not in new:
            changes.append(f"REMOVED: {o['name']} (dropped from extern_companies.json)")
    return changes


MONTH_NAMES = ["", "January", "February", "March", "April", "May", "June",
               "July", "August", "September", "October", "November", "December"]


def format_window(window):
    """Human-readable window: 'August 2026', 'August-October 2026',
    'November 2026 - February 2027'."""
    if not window:
        return "-"
    start, end = window
    sy, sm = int(start[:4]), int(start[5:7])
    ey, em = int(end[:4]), int(end[5:7])
    if (sy, sm) == (ey, em):
        return f"{MONTH_NAMES[sm]} {sy}"
    if sy == ey:
        return f"{MONTH_NAMES[sm]}–{MONTH_NAMES[em]} {sy}"
    return f"{MONTH_NAMES[sm]} {sy} – {MONTH_NAMES[em]} {ey}"


def sort_key(entry, today):
    status = classify(entry, today)
    window = entry.get("window") or ("9999-99", "9999-99")
    return (STATUS_ORDER.index(status) if status in STATUS_ORDER else 99, window[0], entry["name"].lower())


def build_xlsx(entries, today, path=XLSX_FILE):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    STATUS_FILLS = {
        "in_window": "C6EFCE",   # green - Extern expects the window is open now
        "upcoming": "FFEB9C",    # yellow - opens within ~2 months
        "later": "DDEBF7",       # blue - later this cycle
        "passed": "F2F2F2",
        "unknown": "F2F2F2",
        "fetch_failed": "F2F2F2",
        "continuous": "E4DFEC",
    }

    wb = Workbook()
    ws = wb.active
    ws.title = "Watchlist"

    header_font = Font(name="Arial", bold=True, size=10, color="FFFFFF")
    body_font = Font(name="Arial", size=10)
    link_font = Font(name="Arial", size=10, color="0563C1", underline="single")
    header_fill = PatternFill("solid", fgColor="1F4E79")
    wrap = Alignment(wrap_text=True, vertical="top")

    ws["A1"] = (f"Intern Watch - Summer 2027 tech internship calendar - regenerated "
                f"{today.isoformat()} from extern.com company guides. Dates are Extern's "
                f"projections from prior cycles, not confirmed postings. Sorted by how "
                f"soon each window opens.")
    ws["A1"].font = Font(name="Arial", size=9, italic=True, color="595959")
    ws.merge_cells("A1:K1")

    headers = ["Company", "Status", "Expected opening", "Programs & tech tracks",
               "Career page", "Extern guide", "Interview process",
               "Visa / work auth", "Rolling?", "Intern pay",
               "Extern's notes on timing"]
    widths = [22, 24, 24, 44, 32, 10, 46, 36, 28, 28, 44]
    for col, (head, width) in enumerate(zip(headers, widths), start=1):
        cell = ws.cell(row=2, column=col, value=head)
        cell.font = header_font
        cell.fill = header_fill
        ws.column_dimensions[get_column_letter(col)].width = width

    main = [e for e in entries if not e.get("no_formal_program")]
    main.sort(key=lambda e: sort_key(e, today))

    row = 3
    for e in main:
        status = classify(e, today)
        ws.cell(row=row, column=1, value=e["name"]).font = Font(name="Arial", size=10, bold=True)
        scell = ws.cell(row=row, column=2, value=STATUS_LABELS.get(status, status))
        scell.font = body_font
        if status in STATUS_FILLS:
            scell.fill = PatternFill("solid", fgColor=STATUS_FILLS[status])
        ws.cell(row=row, column=3, value=format_window(e.get("window"))).font = body_font
        if e.get("tech_tracks"):
            n_tech = e["tech_tracks"].count(";") + 1
            suffix = (f"  ({n_tech} of {e['tracks_total']} tracks)"
                      if e.get("tracks_total", 0) > n_tech else "")
            programs_value = e["tech_tracks"] + suffix
        else:
            programs_value = e.get("programs") or "See guide"
        pcell = ws.cell(row=row, column=4, value=programs_value)
        pcell.font = body_font
        pcell.alignment = wrap
        urls = e.get("career_urls") or []
        if urls:
            ccell = ws.cell(row=row, column=5, value=re.sub(r"^https?://(www\.)?", "", urls[0])[:60])
            ccell.hyperlink = urls[0]
            ccell.font = link_font
        elif e.get("career_note"):
            ws.cell(row=row, column=5, value=e["career_note"][:60]).font = body_font
        gcell = ws.cell(row=row, column=6, value="guide")
        gcell.hyperlink = e["guide_url"]
        gcell.font = link_font
        notes = e.get("window_text") or e.get("fetch_error") or ""
        for col, value in ((7, e.get("interview") or ""), (8, e.get("visa") or ""),
                           (9, e.get("rolling") or ""), (10, e.get("pay") or ""), (11, notes)):
            cell = ws.cell(row=row, column=col, value=value)
            cell.font = body_font
            cell.alignment = wrap
        row += 1

    ws.freeze_panes = "C3"
    ws.auto_filter.ref = f"A2:K{row - 1}"

    # Companies with no formal internship program go on their own sheet
    # so the main watchlist stays actionable; their career links are
    # still worth a periodic manual look.
    ws2 = wb.create_sheet("No formal program")
    ws2["A1"] = ("Extern reports these companies run no formal internship program (they hire "
                 "ad hoc or full-time only). Not on the main watchlist; check occasionally.")
    ws2["A1"].font = Font(name="Arial", size=9, italic=True, color="595959")
    ws2.merge_cells("A1:D1")
    for col, (head, width) in enumerate(zip(["Company", "How to break in (Extern's words)", "Careers link", "Extern guide"],
                                            [24, 60, 36, 12]), start=1):
        cell = ws2.cell(row=2, column=col, value=head)
        cell.font = header_font
        cell.fill = header_fill
        ws2.column_dimensions[get_column_letter(col)].width = width
    row = 3
    for e in sorted((e for e in entries if e.get("no_formal_program")), key=lambda e: e["name"].lower()):
        ws2.cell(row=row, column=1, value=e["name"]).font = Font(name="Arial", size=10, bold=True)
        cell = ws2.cell(row=row, column=2, value=e.get("window_text") or "")
        cell.font = body_font
        cell.alignment = wrap
        urls = e.get("career_urls") or []
        if urls:
            ccell = ws2.cell(row=row, column=3, value=re.sub(r"^https?://(www\.)?", "", urls[0])[:60])
            ccell.hyperlink = urls[0]
            ccell.font = link_font
        gcell = ws2.cell(row=row, column=4, value="guide")
        gcell.hyperlink = e["guide_url"]
        gcell.font = link_font
        row += 1
    ws2.freeze_panes = "A3"

    # One-click LinkedIn people searches per company - kept as plain
    # search URLs (no scraping/automation of LinkedIn itself), for
    # manual, human-sent outreach.
    from urllib.parse import quote
    ws3 = wb.create_sheet("LinkedIn outreach")
    ws3["A1"] = ("Pre-built LinkedIn people searches. Alumni link: people matching your school + "
                 "the company. Recruiter link: the company's university recruiters. Send messages "
                 "yourself - short, specific, and one at a time.")
    ws3["A1"].font = Font(name="Arial", size=9, italic=True, color="595959")
    ws3.merge_cells("A1:D1")
    for col, (head, width) in enumerate(zip(["Company", "Status", "UF alumni there", "University recruiters"],
                                            [22, 24, 24, 24]), start=1):
        cell = ws3.cell(row=2, column=col, value=head)
        cell.font = header_font
        cell.fill = header_fill
        ws3.column_dimensions[get_column_letter(col)].width = width
    search = "https://www.linkedin.com/search/results/people/?keywords="
    row = 3
    for e in main:
        status = classify(e, today)
        ws3.cell(row=row, column=1, value=e["name"]).font = Font(name="Arial", size=10, bold=True)
        scell = ws3.cell(row=row, column=2, value=STATUS_LABELS.get(status, status))
        scell.font = body_font
        if status in STATUS_FILLS:
            scell.fill = PatternFill("solid", fgColor=STATUS_FILLS[status])
        acell = ws3.cell(row=row, column=3, value="alumni search")
        acell.hyperlink = search + quote(f'"University of Florida" "{e["name"]}"')
        acell.font = link_font
        rcell = ws3.cell(row=row, column=4, value="recruiter search")
        rcell.hyperlink = search + quote(f'"{e["name"]}" "university recruiter"')
        rcell.font = link_font
        row += 1
    ws3.freeze_panes = "A3"

    wb.save(path)


def load_json(path, default):
    if not os.path.exists(path):
        return default
    with open(path) as fh:
        return json.load(fh)


def main():
    config = load_json(CONFIG_FILE, {})
    companies = config.get("companies", [])
    if not companies:
        print(f"No companies in {CONFIG_FILE}.")
        return

    prev = load_json(CALENDAR_FILE, {})
    today = date.today()

    print(f"Scraping {len(companies)} Extern tech guides...")
    entries = scrape_all(companies)

    changes = diff_calendars(prev.get("companies", []), entries)

    # Flag guides Extern newly published (any category) for manual
    # review - never auto-added since many are non-tech.
    known = set(prev.get("known_guide_slugs", []))
    try:
        current_slugs = sitemap_guide_slugs()
    except Exception as exc:  # noqa: BLE001
        print(f"Sitemap check failed: {exc}")
        current_slugs = sorted(known)
    if known:
        fresh = [s for s in current_slugs if s not in known]
        if fresh:
            changes.append(
                "New guides published on extern.com (tech or not - review and add to "
                "extern_companies.json if wanted): " + ", ".join(fresh)
            )

    calendar = {
        "generated_at": datetime.now(timezone.utc).isoformat(),
        "companies": entries,
        "known_guide_slugs": current_slugs,
    }
    with open(CALENDAR_FILE, "w") as fh:
        json.dump(calendar, fh, indent=2)
    build_xlsx(entries, today)
    print(f"Wrote {CALENDAR_FILE} and {XLSX_FILE}.")

    if not prev:
        print("First run - baseline saved, no alert email.")
        return
    if not changes:
        print("No changes since last run.")
        return

    body = "Extern calendar changes:\n\n" + "\n\n".join(changes)
    print(body)
    if os.environ.get("SMTP_HOST"):
        from monitor import send_email
        send_email(f"[intern-watch] Extern calendar: {len(changes)} change(s)", body)
        print("Alert email sent.")


if __name__ == "__main__":
    sys.exit(main())
